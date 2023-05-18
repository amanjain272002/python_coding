import openpyxl
# from openpyxl import load_workbook

book =openpyxl.load_workbook('exs.xlsx')
sheet = book.active
rows = sheet.rows

headers = [cell.value for cell in next(rows)]

max_rows = sheet.max_row

for i in range(1,max_rows+1):
    sheet.cell(row=i, column=4, value='hello')

book.save('exs.xlsx')

